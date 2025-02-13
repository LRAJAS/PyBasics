import openpyxl as xl
from openpyxl.chart import BarChart, LineChart, PieChart, Reference

# Function to create a chart
def create_chart(sheet, chart_type, values, start_col):
    if chart_type == "bar":
        chart = BarChart()
    elif chart_type == "line":
        chart = LineChart()
    elif chart_type == "pie":
        chart = PieChart()
    else:
        print("Invalid chart type! Defaulting to BarChart.")
        chart = BarChart()
    
    chart.add_data(values, titles_from_data=True)
    chart.title = "Data Analysis Chart"
    chart.x_axis.title = "Rows"
    chart.y_axis.title = "Values"
    
    # Insert the chart after the last column with a gap
    start_cell = sheet.cell(row=1, column=start_col + 2).coordinate
    sheet.add_chart(chart, start_cell)

# Function to modify a row or column
def modify_row_or_column(sheet):
    choice = input("Do you want to modify a row or column? (row/column): ").lower()
    index = int(input(f"Enter the {choice} number to modify: "))
    values = input(f"Enter the values separated by commas: ").split(",")

    if choice == "row":
        for col, value in enumerate(values, start=1):
            sheet.cell(row=index, column=col).value = value
    elif choice == "column":
        for row, value in enumerate(values, start=1):
            sheet.cell(row=row, column=index).value = value

# Function to read a row or column
def read_row_or_column(sheet):
    choice = input("Do you want to read a row or column? (row/column): ").lower()
    index = int(input(f"Enter the {choice} number to read: "))
    
    if choice == "row":
        values = [sheet.cell(row=index, column=col).value for col in range(1, sheet.max_column + 1)]
    elif choice == "column":
        values = [sheet.cell(row=row, column=index).value for row in range(1, sheet.max_row + 1)]
    print(f"The {choice} values are: {values}")

# Function to calculate the sum of a row or column
def calculate_sum(sheet):
    choice = input("Do you want to sum a row or column? (row/column): ").lower()
    index = int(input(f"Enter the {choice} number to sum: "))
    
    if choice == "row":
        values = [sheet.cell(row=index, column=col).value for col in range(1, sheet.max_column + 1)]
    elif choice == "column":
        values = [sheet.cell(row=row, column=index).value for row in range(1, sheet.max_row + 1)]
    
    numeric_values = [v for v in values if isinstance(v, (int, float))]
    total = sum(numeric_values)
    print(f"The sum of the {choice} is: {total}")

# Function to perform mathematical operations on a row or column
def perform_operation(sheet):
    choice = input("Do you want to perform operations on a row or column? (row/column): ").lower()
    index = int(input(f"Enter the {choice} number to operate on: "))
    operation = input("Enter the operation (add/sub/mul/div): ").lower()
    operand = float(input("Enter the operand for the operation: "))
    
    if choice == "row":
        for col in range(1, sheet.max_column + 1):
            cell = sheet.cell(row=index, column=col)
            if isinstance(cell.value, (int, float)):
                if operation == "add":
                    cell.value += operand
                elif operation == "sub":
                    cell.value -= operand
                elif operation == "mul":
                    cell.value *= operand
                elif operation == "div" and operand != 0:
                    cell.value /= operand
    elif choice == "column":
        for row in range(1, sheet.max_row + 1):
            cell = sheet.cell(row=row, column=index)
            if isinstance(cell.value, (int, float)):
                if operation == "add":
                    cell.value += operand
                elif operation == "sub":
                    cell.value -= operand
                elif operation == "mul":
                    cell.value *= operand
                elif operation == "div" and operand != 0:
                    cell.value /= operand

# Main program
def main():
    # Get user input for the workbook
    workbook_name = input("Enter the name of the workbook (with .xlsx extension): ")
    
    try:
        # Load the workbook
        wb = xl.load_workbook(workbook_name)
        sheet = wb.active  # Use the active sheet by default
        
        while True:
            # Display menu options
            print("\nMenu Options:")
            print("1. Modify Row/Column")
            print("2. Read Row/Column")
            print("3. Calculate Sum of Row/Column")
            print("4. Perform Mathematical Operation on Row/Column")
            print("5. Add Chart")
            print("6. Exit")
            
            choice = int(input("Enter your choice: "))
            
            if choice == 1:
                modify_row_or_column(sheet)
            elif choice == 2:
                read_row_or_column(sheet)
            elif choice == 3:
                calculate_sum(sheet)
            elif choice == 4:
                perform_operation(sheet)
            elif choice == 5:
                chart_type = input("Enter the chart type (bar/line/pie). Default is 'bar': ").lower() or "bar"
                last_col = sheet.max_column
                last_row = sheet.max_row
                values = Reference(sheet, min_row=1, max_row=last_row, min_col=1, max_col=last_col)
                create_chart(sheet, chart_type, values, last_col)
            elif choice == 6:
                break
            else:
                print("Invalid choice. Please try again.")
        
        # Save the workbook
        wb.save(workbook_name)
        print(f"Workbook updated and saved as '{workbook_name}'.")
    except FileNotFoundError:
        print(f"Error: The file '{workbook_name}' does not exist.")
    except Exception as e:
        print(f"An error occurred: {e}")

# Run the program
if __name__ == "__main__":
    main()
