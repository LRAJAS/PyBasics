import openpyxl as xl
from openpyxl.chart import BarChart, Reference

# Load the workbook and select the sheet
wb = xl.load_workbook('pyproject.xlsx')
sheet = wb['Sheet1']

# Loop through the rows and update the corrected price
for row in range(2, sheet.max_row + 1):  # Assuming data starts from row 2
    cell = sheet.cell(row, 4)  # Column 4 (D)
    if isinstance(cell.value, (int, float)):  # Ensure the value is numeric
        corrected_price = cell.value * 0.9  # Applying a 10% discount
        corrected_price_cell = sheet.cell(row, 5)  # Put the corrected price in column 5 (E)
        corrected_price_cell.value = corrected_price
    else:
        print(f"Skipping row {row} due to non-numeric value in column 4.")

# Create a reference for the chart
values = Reference(sheet, 
                   min_row=2, 
                   max_row=sheet.max_row, 
                   min_col=5,  # Corrected prices are in column 5 (E)
                   max_col=5)

# Create and add the chart
chart = BarChart()
chart.add_data(values, titles_from_data=False)  # Change to False if no headers
chart.title = "Corrected Prices"  # Add a title
chart.x_axis.title = "Item Number"  # Add x-axis label
chart.y_axis.title = "Price"  # Add y-axis label

sheet.add_chart(chart, 'G2')  # Place the chart at G2 to avoid overlap

# Save the workbook
wb.save('pyproject_2.xlsx')
print("Workbook updated and saved as 'pyproject_2.xlsx'.")
